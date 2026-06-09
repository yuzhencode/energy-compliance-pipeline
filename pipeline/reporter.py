"""
pipeline/reporter.py
====================
Renders validated Pydantic model lists into a styled Excel workbook
and optionally uploads it to S3.
"""

from __future__ import annotations

from datetime import date
from pathlib import Path
from typing import List, Optional

import boto3
import openpyxl
import pandas as pd
from botocore.exceptions import BotoCoreError, ClientError
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config import config
from logger import get_logger
from pipeline.models import AccountDetailRow, OfgemSummaryRow

logger = get_logger(__name__)

_HEADER_FILL = PatternFill("solid", fgColor="1F6B8A")
_HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
_THIN_SIDE   = Side(style="thin", color="CCCCCC")
_CELL_BORDER = Border(bottom=_THIN_SIDE)

# Human-readable column names for Excel output
_SUMMARY_RENAME = {
    "avg_debt_no_arrangement_gbp":   "Avg Debt — No Arrangement (£)",
    "avg_debt_with_arrangement_gbp": "Avg Debt — With Arrangement (£)",
    "pct_repaying_via_ppm":          "% Repaying via PPM",
    "accounts_with_debt":            "Accounts with Debt",
    "accounts_no_arrangement":       "Accounts — No Arrangement",
    "total_debt_over_91_days_gbp":   "Total Debt >91 Days (£)",
    "reporting_quarter":             "Reporting Quarter",
    "report_date":                   "Report Date",
}

_DETAIL_RENAME = {
    "account_id":                  "Account ID",
    "customer_name":               "Customer Name",
    "postcode":                    "Postcode",
    "fuel_type":                   "Fuel Type",
    "payment_method":              "Payment Method",
    "debt_amount_gbp":             "Debt Amount (£)",
    "debt_age_days":               "Debt Age (Days)",
    "is_over_91_days":             ">91 Days Flag",
    "account_status":              "Status",
    "has_active_arrangement":      "Has Arrangement",
    "arrangement_date":            "Arrangement Date",
    "arrangement_weekly_rate_gbp": "Weekly Rate (£)",
    "arrangement_plan_weeks":      "Plan (Weeks)",
    "last_switch_type":            "Last Switch Type",
    "last_switch_date":            "Last Switch Date",
    "last_switch_outcome":         "Last Switch Outcome",
    "report_date":                 "Report Date",
}


def _style_sheet(ws, df: pd.DataFrame) -> None:
    for col_idx in range(1, len(df.columns) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill      = _HEADER_FILL
        cell.font      = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = _CELL_BORDER
    for col_idx, col in enumerate(df.columns, 1):
        max_len = max(df[col].astype(str).map(len).max(), len(str(col))) + 4
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len, 44)


class Reporter:
    def __init__(self, output_dir: str = config.output_dir) -> None:
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def _models_to_df(self, rows: list, rename: dict) -> pd.DataFrame:
        df = pd.DataFrame([r.model_dump() for r in rows])
        return df.rename(columns={k: v for k, v in rename.items() if k in df.columns})

    def generate(
        self,
        summary_rows:    List[OfgemSummaryRow],
        detail_rows:     List[AccountDetailRow],
        report_week:     date,
    ) -> Path:
        summary_df = self._models_to_df(summary_rows, _SUMMARY_RENAME)
        detail_df  = self._models_to_df(detail_rows,  _DETAIL_RENAME)

        filename = f"ofgem_report_{report_week.strftime('%Y-%m-%d')}.xlsx"
        filepath = self.output_dir / filename

        with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
            summary_df.to_excel(writer, sheet_name="Ofgem Summary",  index=False)
            detail_df.to_excel(writer,  sheet_name="Account Detail", index=False)

        wb = openpyxl.load_workbook(filepath)
        _style_sheet(wb["Ofgem Summary"],  summary_df)
        _style_sheet(wb["Account Detail"], detail_df)
        wb.save(filepath)

        logger.info(f"Report written → {filepath}")
        return filepath

    def upload_to_s3(self, filepath: Path) -> Optional[str]:
        if not config.aws_bucket:
            logger.info("AWS_S3_BUCKET not set — skipping S3 upload")
            return None
        key = f"reports/{filepath.name}"
        try:
            boto3.client("s3", region_name=config.aws_region).upload_file(
                str(filepath), config.aws_bucket, key
            )
            uri = f"s3://{config.aws_bucket}/{key}"
            logger.info(f"Uploaded → {uri}")
            return uri
        except (BotoCoreError, ClientError) as exc:
            logger.error(f"S3 upload failed: {exc}")
            return None
